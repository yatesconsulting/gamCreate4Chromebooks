#!/usr/bin/python3

from openpyxl import load_workbook # read the excel files
import glob # for looping over all Excel files
import os #  for pipe and environ
import csv # for processing output of gam print cros...
import subprocess # for running gam commands


def doesthisouexistingoogleadmin(ou):
	with os.popen('{} info org "{}" nousers'.format(gamexe, ou)) as pipe:
		a = pipe.readlines()
		lena = len(a)
		if (0 == lena):
			return False
		else:
			return True

def gamcreatecommand(ou):
	return '{} create org "{}"'.format(gamexe, ou)

def createou(ou):
	# print(".{} cmd: {}".format(ou, cmd))
	# return a list of OU create commands for this ou and any missing base OUs (recursivly)
	# sloppy error routines, exit program on fail
	# - if creating OU has no /, then it's at the top level, and probably NOT right
	# - if creating OU, find a good base, then build all the OUs after that base
	if not isinstance(ou,str):
		print("ERROR, I only do one OU at a time, something is terribly wrong.")
		exit()
	if ou[-1] == "/":
		print("ERROR, OUs should not end with trailing slash, fix excel and try again.")
		exit()
	if ('/' in ou.lstrip('/')):
		b = '/'.join((ou.split('/'))[:-1])
		if not (doesthisouexistingoogleadmin(b)):
			r = []
			r.append(gamcreatecommand(ou))
			r.extend(createou(b)) # careful, careful, ...
			return r
		else:
			r = []
			r.append(gamcreatecommand(ou))
			return r
	else:
		if (doesthisouexistingoogleadmin(ou)):
			return []
		else:
			print("ERROR, no new root OU created {}, I just won't do it, I just can't bring myself to belive this is correct.  If it is, please create the root manually at admin.google.com, then try again.".format(ou))
			exit()

def fillglobalvarsfromcodeexcelfile(f):
	# see if there is anything to do, get emails from codesfilename
	# schools = [] # list of school initials to process, returned
	wb = load_workbook(f, read_only=True, data_only=True)
	print ("Gathering which school(s) to process, from {}. Error 404's are OK.".format(f))
	wb = load_workbook(f, read_only=True, data_only=True)
	ws1 = wb.active
	# loop over codes.xlsx for school info
	for r in range(2,50):
		school = ws1["a{}".format(r)].value
		targetou = ws1["b{}".format(r)].value
		emailconf = ws1["c{}".format(r)].value
		notes = ws1["d{}".format(r)].value
		location = ws1["e{}".format(r)].value
		destiny = ws1["f{}".format(r)].value
		if (isinstance(emailconf,str) and isinstance(targetou,str) and isinstance(school,str) and school[:2].upper() != "EX" and emailconf[:3].upper() != "NOT"):
			school = school.strip()
			# schools.append(school)
			targetou = targetou.strip()
			emailconf = emailconf.strip()
			if (isinstance(notes,str) ):
				school2notes[school] = notes.strip()
			else:
				school2notes[school] = ""
			if (isinstance(location,str) ):
				school2location[school] = location.strip()
			else:
				school2location[school] = ""
			
			destinypartofmessage = ""
			if (destiny == "Yes"):
				school2destiny[school] =  True
				destinypartofmessage = " and added to Destiny file"
			else:
				school2destiny[school] =  False
			school2ou[school] = targetou
			school2email[school] = emailconf
			ougood = doesthisouexistingoogleadmin(targetou)
			if ougood:
				newouornot = ""
			else:
				newouornot = "a NEW ou "
				# check up one (and up one...) until you find a good root, build all needed OUs
				cmd.extend(createou(targetou))
			print ("*** {} Chromebooks in / by {} will be processed into {}{}{}.".format(school, emailconf, newouornot, targetou, destinypartofmessage))
	wb.close()
	# return schools

def fillglobalvarsfromsearchofexcelfiles(school):
	# work from current directory, open each excel file
	for f in glob.glob("*{}*.xls*".format(school)):
		# skip any temp files
		if f[0] == "~":
			continue
		print ("Adding lookups for serial/tags from {}".format(f))
		wb = load_workbook(f, read_only=True, data_only=True)
		for s in wb.sheetnames:
			# ~ print("worksheet {}".format(s))
			ws1 = wb[s]
			for r in range(2,200):
				tag = ws1["b{}".format(r)].value
				serial = ws1["c{}".format(r)].value
				desc = ws1["d{}".format(r)].value
				school = ws1["e{}".format(r)].value
				room = ws1["f{}".format(r)].value
				fulltag = "{}CB-{}".format(school,tag)
				if (isinstance(serial,str) and serial > "" and serial not in ("No Number","None","Building","BLDG") and isinstance(school,str)): 
					# ERROR, dup serial on line 13: BuildingCB-Tag # <> BLDGCB-Tag # FIX TODO

					serial = serial.strip()
					school = school.strip()
					if serial in sntotag.keys():
						# just ignore if same, otherwise red flag
						if sntotag[serial] != fulltag:
							print ("ERROR, dup serial on line {}: {} <> {}".format(r, sntotag[serial], fulltag))
							# probably do something else here, this could be a real problem TODO
							# sntotag[serial] = 'dups {} {}'.format(sntotag[serial].replace("dups ",""),fulltag)
					else:
						sntotag[serial] = fulltag
						sntoschool[serial] = school
		wb.close()
		# ~ for a in sntotag.keys():
			# ~ print ("key {} = {}".format(a,sntotag[a]))

def gamcroscheck():
	# ~ Now loop over all the OU / CrOS devices and add them to the list, if appropriate
	with os.popen('{} print cros limit_to_ou / fields deviceId,serialNumber,status,lastSync,annotatedUser,annotatedLocation,annotatedAssetId,lastEnrollmentTime,orgUnitPath,notes'.format(gamexe)) as pipe:
		reader = csv.DictReader(pipe)
		for row in reader:
			# ~ if this chromebook sn -> connie's file for school and tag -> codes for owners email
			# ~ If it is a good one to register, do it
			sn = row['serialNumber']
			cbuser = row['annotatedUser']
			status = row['status']
			if (status == "ACTIVE" and sn in sntoschool.keys() and sntoschool[sn] in school2email.keys() and school2email[sntoschool[sn]] == cbuser):
				school = sntoschool[sn]
				deviceid = row['deviceId']
				loc = school2location[school]
				mname = sntotag[sn]
				notes = school2notes[school]
				destou = school2ou[school]		
				cmd.append('{} update cros query:id:{} notes "{}" ou "{}" assetid "{}" location "{}"'.format(gamexe,sn,notes,destou,mname,loc))
				# print(cmd[-1])
				try:
					if (school2destiny[school]):
						# add to destiny CSV also
						s2dcsv.append("{},{},{},{}".format(school,sn,mname,notes))
				except:
					pass

def warnthenruncmd():
	print("--- Here's what I'm really going to do if you say yes: ---")
	for a in cmd:
		print (a)

	l = len(cmd)
	if l > 0:
		sifplural = ""
		if l > 1:
			sifplural = "s"
		doit = input("You sure you want to run {} command{} [Y,n]? ".format(len(cmd),sifplural))
		if (doit.strip() == "" or doit.lower().strip()[0] == "y"):
			print ("Ok, this will take a few minutes...")
			for c in cmd:
				# execute c and display any errors to screen
				print(c)
				r = subprocess.run(c, capture_output=True)
				if (r.stderr):
					print ("  ERROR -- {}".format(r.stderr))
			# create Destiny CSV if s2dcsv > 1 value
			if len(s2dcsv) > 1:
				user = os.environ['USERNAME']
				filename = "{}DestinyChromebooksFile.csv".format(user)
				# drop header line if appending to the file
				if os.path.exists(filename):
					del s2dcsv[0]
					destinyfileaction = 'appended'
				else:
					destinyfileaction = 'created'
				f = open(filename, "a")  
				for l in s2dcsv:
					f.write("{}\n".format(l))
				print ("filename {} {} for Destiny.".format(filename, destinyfileaction))
				f.close()
		else:
			print("Maybe next time, thanks.")
	else:
		print("Sorry, but I couldn't find anything to do.")

if __name__ == "__main__":
	codesfilename = "codes.xlsx"
	school2ou = {}
	school2notes = {}
	school2email = {}
	school2location = {}
	school2destiny = {}
	s2dcsv = ['SCHOOL,BARCODE,CF1:NAME,CF2:NOTES']
	gamexe = "gam"

	# from Connie's file(s), s/n to tag# and school
	# edit (14,44) if you use a different range of rows, around line  117
	sntotag = {}
	sntoschool = {}
	cmd = []
	fillglobalvarsfromcodeexcelfile(codesfilename)
	for s in school2ou.keys():
		if s in ['PEC','PES']:
			fillglobalvarsfromsearchofexcelfiles('pe[cs]')
		else:
			fillglobalvarsfromsearchofexcelfiles(s)
	
	cmd = sorted(list(set(cmd)))
	gamcroscheck()
	warnthenruncmd()
