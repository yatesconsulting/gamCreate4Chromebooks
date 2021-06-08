#!/usr/bin/python3

from openpyxl import load_workbook # read the excel files
import glob # for looping over all Excel files
import os #  for pipe and environ
import csv # for processing output of gam print cros...
import subprocess # for running gam commands
import sys

def searchfile(needle, fnames):
	needle = "{}".format(needle).lower()
	for f in glob.glob(fnames):
		if ".xls" not in f:
			continue
		# print("Searching {}".format(f))
		try:
			wb = load_workbook(f, read_only=True, data_only=True)
			for s in wb.sheetnames:
				tagbeenblank = 0
				ws1 = wb[s]
				# print ("New Sheet found: {}".format(s))
				for r in range(2,200):
					tag = "{} ".format(ws1["b{}".format(r)].value).lower()
					serial = "{} ".format(ws1["c{}".format(r)].value).lower()
					desc = "{} ".format(ws1["d{}".format(r)].value).lower()
					school = "{} ".format(ws1["e{}".format(r)].value).lower()
					room = "{} ".format(ws1["f{}".format(r)].value).lower()
					# print ("needle;;;{};;; in? {}{}{}{}{}".format(needle, tag, serial, desc, school, room))
					if (needle in "{}{}{}{}{}".format(tag, serial, desc, school, room)): # fix to allow partials of each element
						print ("{}/{}: {},{},{},{},{}".format(f, s, tag,serial,desc,school,room))
					if tag == "None" and r > 20:
						# not isinstance(tag,str):
						if tagbeenblank < 5:
							tagbeenblank += 1
						else:
							break
			wb.close()
		except:
			print("Skipping {} because {}".format(f, sys.exc_info()[0]))
		

def getoptions(needle = "", fname = ""):
	# for a in (needle,fname):
	# 	print ("found something: {}".format(a))
	if needle > '' and fname > '':
		return [needle, fname]
	elif (len(sys.argv) > 2):
		needle = sys.argv[1]
		fname = sys.argv[2]
	elif (len(sys.argv) == 2):
		needle = sys.argv[1]
		fname = '*.xls*'
	else:
		print("Enter the search term: ", end="")
		needle = input().strip()
		print("What file(s) do you want me to search? [*.xls*] ", end="")
		fname = input().strip()
		if (fname == ""):
			fname = "*.xls*"
	return [needle, fname]

if __name__ == "__main__":
	# excelgrep blah *files*
	[needle, fnames] = getoptions()
	searchfile(needle, fnames)