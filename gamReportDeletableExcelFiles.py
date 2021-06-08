#!/usr/bin/python3

from openpyxl import load_workbook # read the excel files
import glob # for looping over all Excel files
import os #  for pipe and environ
import csv # for processing output of gam print cros...
import sys
import subprocess # for running gam commands

# def schoolfromou(ou):
# 	parts = ou.split('/')
# 	for a in range(1,len(parts)+1):
# 		p = parts[-a]
# 		if p in validschools:
# 			return p
# 	return ""

def checkthisfile(f, found):
	if f[0] == "~" or f == codesfilename:
		return "Skipping file {}".format(f)
	else:
		# print ("Checking on file {}".format(f))
		try:
			wb = load_workbook(f, read_only=True, data_only=True)
			for s in wb.sheetnames:
				tagbeenblank = 0
				# print("worksheet {}".format(s))
				ws1 = wb[s]
				for r in range(2,300):
					serial = str(ws1["c{}".format(r)].value)
					desc = str(ws1["d{}".format(r)].value).lower().replace(' ','')
					school = str(ws1["e{}".format(r)].value)
					# print("line {}, serial {}, desc {}, school {}".format(r, serial, desc, school))
					if (serial is not None and serial and serial[:4] != "None" and 'hromebook' in desc):
						serial = serial.strip()
						if serial not in found:
							return "Can't delete {}  one, I found {} on worksheet {}".format(f, serial, f , s)
						# else:
						# 	print("found {} in chromebook list".format(serial))
					
					if ((serial is None or serial == "None") and r > 20):
						# not isinstance(tag,str):
						if tagbeenblank < 5:
							tagbeenblank += 1
						else:
							break
			wb.close()
			return 'Ok to del "{}"'.format(f)
		except:
			print("Skipping {} because {}".format(f, sys.exc_info()[0]))

def getcbsns():
	# ~ return list of all CrOS device s/n's not including those in /
	found = []
	# with os.popen('{} print cros limit_to_ou "/STUDENTS/ELEMENTARY/ECE/Replacements/" fields serialNumber,status,orgUnitPath'.format(gamexe)) as pipe:
	with os.popen('{} print cros fields serialNumber,status,orgUnitPath'.format(gamexe)) as pipe:
		reader = csv.DictReader(pipe)
		for row in reader:
			# ~ if this chromebook sn -> connie's file for school and tag -> codes for owners email
			# ~ If it is a good one to register, do it
			# status = row['status']
			ou = row['orgUnitPath']
			sn = row['serialNumber']
			if (ou != "/"):
				# ou = row['orgUnitPath']
				# school = schoolfromou(ou)
				# if (school):
				# 	deviceid = row['deviceId']
				# 	loc = school2location[school]
				# 	mname = sntotag[sn]
				found.append(sn)
	return found

if __name__ == "__main__":
	gamexe = "gam"
	codesfilename = "codes.xlsx"

	# from Connie's file(s), s/n to tag# and school
	# edit (14,44) if you use a different range of rows, around line  117
	validschools = ('BES','BMS','MVM','DNE','MMS','EGP','ECE','NLE',\
		'MHE','MAE','MON','SUN','SMS','PES','GHS','UHS','ECHS','RHS','VVE','WAE')

	# work from current directory, open each excel file
	InGSuiteSNs = getcbsns()

	f = open("OUTTEST.csv", "a")  
	for l in InGSuiteSNs:
		f.write("{}\n".format(l))
	f.close()

	for f in glob.glob("*.xls*"):
		print(checkthisfile(f, InGSuiteSNs))

	